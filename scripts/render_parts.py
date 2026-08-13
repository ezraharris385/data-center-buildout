#!/usr/bin/env python3
"""render_parts.py — authoritative render-detail data for Data Center Buildout.

Generates BOTH:
  data/render_parts.xlsx  (the editable source of truth — open it, fix a dimension,
                           re-run this script with --from-xlsx to update the app)
  data/render_parts.json  (what the app actually loads)

Schema (sheet "Parts", one row per sub-part):
  Component_ID  catalog SKU this part belongs to
  Part          human name
  Shape         box | cyl | cylH | torus | fan | vpanel
  Cx/Cy/Cz_mm   part CENTER position: X across width, Y height above ground, Z depth
                (+Z = component front). Whole component centered at origin, facing +Z.
  W/D/H_mm      box: width/depth/height. vpanel: length/thickness/panel height.
  Dia_mm        cyl, cylH, fan, torus (ring diameter)
  Len_mm        cyl height, cylH length, torus tube diameter
  Count         array copies (default 1)
  Axis          array axis x|y|z
  Spacing_mm    array pitch
  RotX/Y/Z_deg  rotation applied to each copy
  Material      palette key in materials.js (or hex like #aabbcc)
  Spin          1 = animated fan hub
  Blades        fan blade count
  Notes         source / rationale

All dims mm. Sources: manufacturer datasheets & installation drawings (recall-grade,
same confidence class as the base catalog).
"""
import json, os, sys

OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')
COLS = ['Component_ID','Part','Shape','Cx_mm','Cy_mm','Cz_mm','W_mm','D_mm','H_mm',
        'Dia_mm','Len_mm','Count','Axis','Spacing_mm','RotX_deg','RotY_deg','RotZ_deg',
        'Material','Spin','Blades','Notes']

def P(cid, part, shape, cx=0, cy=0, cz=0, w=0, d=0, h=0, dia=0, ln=0, count=1,
      axis='x', spacing=0, rx=0, ry=0, rz=0, mat='switchgear', spin=0, blades=6, notes=''):
    return dict(zip(COLS, [cid, part, shape, cx, cy, cz, w, d, h, dia, ln, count,
                           axis, spacing, rx, ry, rz, mat, spin, blades, notes]))

PARTS = []

# ============================================================ MEC-005
# Güntner GFD V-shape dry cooler, 1 MW class. 12000 x 2275 x 2850.
PARTS += [
    P('MEC-005','Base frame rails','box', 0,100,0, 12000,2200,200, mat='gensetDark',
      notes='Galvanized frame, twin longitudinal rails'),
    P('MEC-005','V-coil bank left','vpanel', 0,1450,-620, 11700,80,1950, rx=-26, mat='chillerCoil',
      notes='Microchannel coil, ~26 deg from vertical, full length'),
    P('MEC-005','V-coil bank right','vpanel', 0,1450,620, 11700,80,1950, rx=26, mat='chillerCoil'),
    P('MEC-005','End panel A','box', -5960,1500,0, 80,2275,2000, mat='chiller'),
    P('MEC-005','End panel B','box', 5960,1500,0, 80,2275,2000, mat='chiller'),
    P('MEC-005','Fan deck','box', 0,2480,0, 12000,2100,70, mat='chiller'),
    P('MEC-005','EC fan array','fan', -5075,2600,0, dia=910, count=8, axis='x', spacing=1450,
      mat='fanBlade', spin=1, blades=7, notes='8x EC axial fans, ~910 mm impeller'),
    P('MEC-005','Supply header','cylH', 0,600,1120, dia=168, ln=11200, mat='pipeSupply',
      notes='DN150 glycol supply header'),
    P('MEC-005','Return header','cylH', 0,950,1120, dia=168, ln=11200, mat='pipeReturn'),
]

# ============================================================ MEC-001
# Carrier 30XA-1002 air-cooled screw chiller. 11937 x 2253 x 2297.
PARTS += [
    P('MEC-001','Base rails','box', 0,90,0, 11900,2200,180, mat='gensetDark'),
    P('MEC-001','V-coil left','vpanel', -1800,1300,-560, 8200,70,1750, rx=-30, mat='chillerCoil',
      notes='Condenser coils occupy ~8.2 m of length; compressor end clear'),
    P('MEC-001','V-coil right','vpanel', -1800,1300,560, 8200,70,1750, rx=30, mat='chillerCoil'),
    P('MEC-001','Condenser fans','fan', -5050,2230,0, dia=800, count=6, axis='x', spacing=1300,
      mat='fanBlade', spin=1, blades=6, notes='6 fans over coil section'),
    P('MEC-001','Control panel','box', 5300,1050,0, 900,2200,1900, mat='pdu',
      notes='Unit-mounted control/starter cabinet at end'),
    P('MEC-001','Screw compressor A','cylH', 4200,750,-500, dia=520, ln=2100, mat='gensetDark',
      notes='Twin 06T-class screw compressors'),
    P('MEC-001','Screw compressor B','cylH', 4200,750,500, dia=520, ln=2100, mat='gensetDark'),
    P('MEC-001','Oil separator A','cyl', 3300,800,-500, dia=420, ln=1300, mat='tank'),
    P('MEC-001','Oil separator B','cyl', 3300,800,500, dia=420, ln=1300, mat='tank'),
    P('MEC-001','Evaporator shell','cylH', 1000,650,0, dia=900, ln=6500, mat='tank',
      notes='Flooded evaporator barrel under coils'),
    P('MEC-001','CHW supply stub','cylH', -2200,650,1180, dia=219, ln=400, ry=90, mat='pipeSupply'),
    P('MEC-001','CHW return stub','cylH', -3000,650,1180, dia=219, ln=400, ry=90, mat='pipeReturn'),
]

# ============================================================ MEC-002
# Vertiv Liebert AFC free-cooling chiller (1 MW). 10750 x 2250 x 2540. Enclosed.
PARTS += [
    P('MEC-002','Base frame','box', 0,90,0, 10700,2200,180, mat='gensetDark'),
    P('MEC-002','Body shell','box', 0,1150,0, 10700,2250,1950, mat='chiller'),
    P('MEC-002','Louver band left','box', 0,1150,-1140, 10500,50,1700, mat='chillerCoil'),
    P('MEC-002','Louver band right','box', 0,1150,1140, 10500,50,1700, mat='chillerCoil'),
    P('MEC-002','Fan array','fan', -4450,2320,0, dia=800, count=6, axis='x', spacing=1780,
      mat='fanBlade', spin=1, blades=7),
    P('MEC-002','Electrical panel','box', 5100,1100,0, 500,2100,1800, mat='pdu'),
]

# ============================================================ MEC-004
# BAC Series 3000 cooling tower cell, 12x18 ft. 5486 x 3658 x 4900.
PARTS += [
    P('MEC-004','Cold water basin','box', 0,350,0, 5486,3658,700, mat='gensetDark'),
    P('MEC-004','Casing','box', 0,2100,0, 5400,3600,2800, mat='coolingTower'),
    P('MEC-004','Air inlet louvers A','box', 0,1500,-1790, 5300,60,1800, mat='chillerCoil'),
    P('MEC-004','Air inlet louvers B','box', 0,1500,1790, 5300,60,1800, mat='chillerCoil'),
    P('MEC-004','Fan cylinder','cyl', 0,3900,0, dia=2700, ln=1000, mat='coolingTower',
      notes='FRP fan stack'),
    P('MEC-004','Axial fan','fan', 0,3950,0, dia=2400, mat='fanBlade', spin=1, blades=7,
      notes='Single low-speed axial fan'),
    P('MEC-004','Fan motor','box', 900,4550,0, 600,500,450, mat='gensetDark'),
    P('MEC-004','Hot water inlet','cylH', 0,2950,1900, dia=273, ln=600, ry=90, mat='pipeReturn'),
]

# ============================================================ MEC-007
# Chilled-water TES tank ~1 MG. 15240 dia x 21336. (Rendered at 55% scale in scene.)
PARTS += [
    P('MEC-007','Tank shell','cyl', 0,10000,0, dia=15240, ln=20000, mat='tank',
      notes='Welded steel, field-erected'),
    P('MEC-007','Roof cone','cyl', 0,20600,0, dia=15300, ln=1200, mat='tank'),
    P('MEC-007','Ring stiffener','torus', 0,4000,0, dia=15400, ln=250, count=4, axis='y',
      spacing=4500, rx=90, mat='gensetDark', notes='External stiffening rings'),
    P('MEC-007','Access ladder','box', 7700,10500,0, 150,450,21000, mat='gensetDark'),
    P('MEC-007','Diffuser pipe','cylH', 0,1200,7900, dia=500, ln=1500, ry=90, mat='pipeSupply'),
]

# ============================================================ FUE-001
# UL-2085 protected AST, 20,000 gal. 10973 x 3048 x 3400.
PARTS += [
    P('FUE-001','Tank shell','cylH', 0,1800,0, dia=3000, ln=10400, mat='fuelTank'),
    P('FUE-001','Saddle A','box', -3000,800,0, 350,3048,1600, mat='gensetDark'),
    P('FUE-001','Saddle B','box', 3000,800,0, 350,3048,1600, mat='gensetDark'),
    P('FUE-001','Manway','cyl', 0,3450,0, dia=600, ln=350, mat='gensetDark'),
    P('FUE-001','Vent stack','cyl', 4600,4000,0, dia=100, ln=1400, mat='gensetDark'),
    P('FUE-001','Fill cabinet','box', -4900,1100,900, 600,500,900, mat='pdu',
      notes='Fill/spill containment cabinet at end'),
    P('FUE-001','Level gauge','box', -5550,1800,0, 120,300,500, mat='pdu'),
]

# ============================================================ BKP-003
# Enclosed 3 MW genset, 40 ft-class. 12192 x 2438 x 3960.
PARTS += [
    P('BKP-003','Sub-base fuel tank','box', 0,300,0, 12000,2438,600, mat='gensetDark',
      notes='Full-length UL-142 sub-base tank'),
    P('BKP-003','Enclosure','box', -500,2050,0, 11000,2438,2900, mat='gensetEnclosure'),
    P('BKP-003','Intake louver wall','box', -6030,2050,0, 60,2300,2300, mat='chillerCoil',
      notes='Cool air intake at alternator end'),
    P('BKP-003','Side louvers','box', -2500,2300,1230, 3200,40,1400, mat='chillerCoil'),
    P('BKP-003','Radiator discharge hood','box', 5550,2050,0, 1100,2438,2900, mat='gensetDark'),
    P('BKP-003','Radiator fan','fan', 6130,2150,0, dia=1600, rz=90, mat='fanBlade', spin=1,
      blades=5, notes='Horizontal-discharge radiator fan, faces away from building'),
    P('BKP-003','Exhaust stack','cyl', -3400,4750,500, dia=350, ln=1900, mat='gensetDark'),
    P('BKP-003','Rain cap','box', -3400,5780,500, 550,550,120, mat='gensetDark'),
    P('BKP-003','Silencer','cylH', -1500,4150,500, dia=800, ln=3500, mat='tank',
      notes='Critical-grade silencer on roof'),
    P('BKP-003','Access doors','box', -2500,1800,-1230, 2600,30,1800, count=2, axis='x',
      spacing=3400, mat='gensetDark'),
    P('BKP-003','Status beacon','box', -5800,3650,0, 120,120,120, mat='ledAmber'),
]

# ============================================================ BKP-001
# Caterpillar 3516B open genset, 2000 kW. 6100 x 2170 x 2620.
PARTS += [
    P('BKP-001','Skid base','box', 0,140,0, 6100,2170,280, mat='gensetDark'),
    P('BKP-001','Engine block','box', -700,1150,0, 2900,1450,1500, mat='genset',
      notes='V16 block, CAT yellow'),
    P('BKP-001','Valve covers','box', -700,2020,0, 2700,1250,260, mat='genset'),
    P('BKP-001','Turbo + aftercooler','box', -2000,2100,0, 700,900,500, mat='gensetDark'),
    P('BKP-001','Alternator','cylH', 1750,1050,0, dia=1350, ln=1700, mat='gensetDark'),
    P('BKP-001','Terminal box','box', 2500,1900,0, 700,700,500, mat='pdu'),
    P('BKP-001','Radiator','box', -2870,1250,0, 360,2170,1950, mat='chillerCoil'),
    P('BKP-001','Radiator fan','fan', -2600,1250,0, dia=1500, rz=90, mat='fanBlade', spin=1, blades=6),
    P('BKP-001','Exhaust flex','cyl', -1200,2600,400, dia=300, ln=700, mat='tank'),
    P('BKP-001','Air cleaner','cylH', 0,2450,-600, dia=450, ln=1000, mat='gensetDark'),
]

# ============================================================ BKP-002
# Cummins QSK95, 3500 kW. 6800 x 2440 x 2900. Same anatomy, larger.
PARTS += [
    P('BKP-002','Skid base','box', 0,150,0, 6800,2440,300, mat='gensetDark'),
    P('BKP-002','Engine block','box', -800,1300,0, 3200,1600,1700, mat='#3a5e46',
      notes='QSK95 V16, Cummins green-black'),
    P('BKP-002','Valve covers','box', -800,2280,0, 3000,1400,280, mat='#3a5e46'),
    P('BKP-002','Turbos','box', -2300,2350,0, 800,1000,550, mat='gensetDark'),
    P('BKP-002','Alternator','cylH', 1950,1150,0, dia=1500, ln=1900, mat='gensetDark'),
    P('BKP-002','Terminal box','box', 2800,2100,0, 800,800,550, mat='pdu'),
    P('BKP-002','Radiator','box', -3200,1400,0, 400,2440,2200, mat='chillerCoil'),
    P('BKP-002','Radiator fan','fan', -2900,1400,0, dia=1700, rz=90, mat='fanBlade', spin=1, blades=6),
    P('BKP-002','Exhaust flex','cyl', -1400,2900,450, dia=350, ln=800, mat='tank'),
]

# ============================================================ ELC-008
# 2500 kVA dry-type/cast-resin transformer. 2500 x 1500 x 2450.
PARTS += [
    P('ELC-008','Core & coil tank','box', 0,1000,0, 1450,1350,1850, mat='transformer'),
    P('ELC-008','Radiator bank A','box', -960,950,-390, 380,70,1500, count=6, axis='z',
      spacing=190, mat='transformerFin', notes='Panel radiators, 6 per side'),
    P('ELC-008','Radiator bank B','box', 960,950,-390, 380,70,1500, count=6, axis='z',
      spacing=190, mat='transformerFin'),
    P('ELC-008','Conservator','cylH', 0,2280,-350, dia=430, ln=1350, mat='tank',
      notes='Oil conservator drum on top'),
    P('ELC-008','HV bushings','cyl', -500,2500,150, dia=130, ln=550, count=3, axis='x',
      spacing=500, mat='#e8e0cc', notes='Porcelain, 3-phase'),
    P('ELC-008','LV bushings','cyl', -350,2320,450, dia=90, ln=350, count=3, axis='x',
      spacing=350, mat='#e8e0cc'),
    P('ELC-008','Cooling fan','fan', -960,300,0, dia=500, rx=90, mat='fanBlade', spin=1, blades=5,
      notes='Forced-air fan under radiator bank'),
]

# ============================================================ LCL-002
# Vertiv Liebert XDU 1350 row CDU. 750 x 1200 x 2000.
PARTS += [
    P('LCL-002','Cabinet','box', 0,1000,0, 750,1200,2000, mat='cdu'),
    P('LCL-002','Front door','box', 0,1050,615, 700,25,1800, mat='crahDark'),
    P('LCL-002','HMI display','box', 0,1620,632, 300,15,190, mat='screenDark'),
    P('LCL-002','Lower service grille','box', 0,330,618, 690,18,480, mat='chillerCoil'),
    P('LCL-002','TCS supply (to racks)','cylH', -190,420,-680, dia=114, ln=380, ry=90, mat='pipeSupply',
      notes='Blind-mate secondary loop supply'),
    P('LCL-002','TCS return (from racks)','cylH', 190,420,-680, dia=114, ln=380, ry=90, mat='pipeReturn'),
    P('LCL-002','FWS supply (facility)','cylH', -190,820,-680, dia=141, ln=380, ry=90, mat='pipeSupply'),
    P('LCL-002','FWS return (facility)','cylH', 190,820,-680, dia=141, ln=380, ry=90, mat='pipeReturn'),
    P('LCL-002','Status LED','box', 260,1850,632, 40,10,40, mat='ledBlue'),
]

# ============================================================ ACL-002
# Vertiv Liebert CW146 perimeter CRAH. 2921 x 889 x 1969.
PARTS += [
    P('ACL-002','Cabinet','box', 0,985,0, 2921,889,1969, mat='crah'),
    P('ACL-002','Discharge grille','box', 0,530,455, 2800,20,860, mat='crahDark',
      notes='Downflow front discharge'),
    P('ACL-002','Filter access band','box', 0,1580,455, 2800,20,540, mat='crahDark'),
    P('ACL-002','HMI','box', 1180,1560,465, 300,15,200, mat='screenDark'),
    P('ACL-002','CHW supply','cylH', -1000,300,-500, dia=141, ln=350, ry=90, mat='pipeSupply'),
    P('ACL-002','CHW return','cylH', -700,300,-500, dia=141, ln=350, ry=90, mat='pipeReturn'),
]

# ============================================================ ELC-001
# Vertiv EXL S1 1200 kVA UPS lineup. 3300 x 991 x 2134.
PARTS += [
    P('ELC-001','Lineup body','box', 0,1067,0, 3300,991,2134, mat='upsBody'),
    P('ELC-001','Cabinet doors','box', -1237,1040,510, 780,25,1980, count=4, axis='x',
      spacing=825, mat='upsAccent', notes='4 sections: rectifier/inverter/static switch/battery interface'),
    P('ELC-001','HMI touchscreen','box', -1237,1650,528, 340,15,220, mat='screenDark'),
    P('ELC-001','Top cable duct','box', 0,2210,0, 3300,600,150, mat='upsAccent'),
    P('ELC-001','Vent band','box', 0,180,510, 3200,15,220, mat='chillerCoil'),
]

# ============================================================ ELC-002
# Schneider Galaxy VX 1500 kVA. 5600 x 970 x 1970.
PARTS += [
    P('ELC-002','Lineup body','box', 0,985,0, 5600,970,1970, mat='upsBody'),
    P('ELC-002','Cabinet doors','box', -2330,960,500, 880,25,1820, count=6, axis='x',
      spacing=932, mat='upsAccent', notes='6-frame lineup'),
    P('ELC-002','HMI','box', -2330,1560,518, 360,15,230, mat='screenDark'),
    P('ELC-002','Vent band','box', 0,170,500, 5500,15,200, mat='chillerCoil'),
]

HEADER_NOTES = [
    ['Data Center Buildout — render detail pack'],
    ['Edit any dimension below, then run: python3 scripts/render_parts.py --from-xlsx'],
    ['All dimensions in mm. Coordinates are part centers; component origin at ground center, front = +Z.'],
    [''],
]

def write_xlsx(rows, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Parts'
    ws.append(COLS)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1F3B4D')
    for r in rows:
        ws.append([r[c] for c in COLS])
    ws.freeze_panes = 'A2'
    for i, w in enumerate([13,24,8,8,8,8,8,8,8,8,8,7,6,11,9,9,9,14,5,7,52], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    # README sheet
    rs = wb.create_sheet('README')
    for line in [
        ['Render detail pack — how to use'], [''],
        ['1. Every row is one sub-part of a catalog component (Component_ID).'],
        ['2. Shapes: box (W/D/H), cyl (vertical, Dia/Len), cylH (horizontal along X, Dia/Len),'],
        ['   torus (Dia ring / Len tube), fan (Dia, animated if Spin=1), vpanel (tilted panel, use RotX).'],
        ['3. Cx/Cy/Cz are the part CENTER in mm. Y is height above ground. +Z is the component front.'],
        ['4. Count/Axis/Spacing arrays a part (e.g. 8 fans along X at 1450 mm pitch).'],
        ['5. Material: palette key from js/materials.js, or a #rrggbb hex.'],
        ['6. After editing: python3 scripts/render_parts.py --from-xlsx   (regenerates the JSON the app loads)'],
    ]:
        rs.append(line)
    wb.save(path)

def write_json(rows, path):
    by_id = {}
    for r in rows:
        by_id.setdefault(r['Component_ID'], []).append(r)
    with open(path, 'w') as f:
        json.dump(by_id, f, indent=1)

def read_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Parts']
    rows = []
    hdr = [c.value for c in ws[1]]
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw[0]:
            continue
        d = dict(zip(hdr, raw))
        for k in COLS:
            if d.get(k) is None:
                d[k] = 0 if k not in ('Component_ID','Part','Shape','Axis','Material','Notes') else ''
        rows.append(d)
    return rows

if __name__ == '__main__':
    os.makedirs(OUT_DIR, exist_ok=True)
    xlsx = os.path.join(OUT_DIR, 'render_parts.xlsx')
    js = os.path.join(OUT_DIR, 'render_parts.json')
    if '--from-xlsx' in sys.argv:
        rows = read_xlsx(xlsx)
        write_json(rows, js)
        print(f'json regenerated from xlsx: {len(rows)} parts')
    else:
        write_xlsx(PARTS, xlsx)
        write_json(PARTS, js)
        ids = sorted({p["Component_ID"] for p in PARTS})
        print(f'wrote {len(PARTS)} parts for {len(ids)} components: {", ".join(ids)}')
